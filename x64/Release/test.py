import pandas as pd
from openpyxl.styles import PatternFill
# from openpyxl.utils.dataframe_to_excel import dataframe_to_excel
from openpyxl import load_workbook

pre_result_txt = "result.txt"
image_dir = "test/images"
gt_json_dir = "test/jsons"


class ResultDecoder(object):
    def __init__(self,txt_path,excle_path):
        self.txt_path = txt_path
        self.txt_lines = []        
        self.result = {}
        self.result_path = excle_path

        self.paths = []
        self.detect_nums = []
        self.decode_nums = []
        self.decode_rates = []
        self.spend_times = []
        self.scene = []
    
    def read(self):
        with open(self.txt_path, 'r') as f:
            self.txt_lines = f.readlines()
            for line in self.txt_lines:
                file_path,barcode_info,latency = line.split(" ")
                if "common" in file_path:
                    self.paths.append(file_path)
                    self.scene.append("common")
                    self.spend_times.append(latency)
                    barcodes = barcode_info.split("^")
                    detect_num = len(barcodes)
                    self.detect_nums.append(detect_num)
                    decode_num = 0
                    for b in barcodes:
                        decode_info = b.split(",")[0]
                        if decode_info != "":
                            decode_num += 1
                    self.decode_nums.append(decode_num)
                    decode_rate = decode_num/detect_num
                    self.decode_rates.append(decode_rate)
            if sum(self.detect_nums):
                self.paths.append("-")
                self.detect_nums.append(sum(self.detect_nums))
                self.decode_nums.append(sum(self.decode_nums))
                self.decode_rates.append(sum(self.decode_nums)/sum(self.detect_nums))
                self.spend_times.append("-")
                self.scene.append("common total")


            # dark scene total
            self.dark_paths = []
            self.dark_detect_nums = []
            self.dark_decode_nums = []
            self.dark_decode_rates = []
            self.dark_spend_times = []
            self.dark_scene = []
            for line in self.txt_lines:
                file_path,barcode_info,latency = line.split(" ")
                if "dark" in file_path:
                    self.dark_paths.append(file_path)
                    self.dark_scene.append("dark")
                    self.dark_spend_times.append(latency)
                    barcodes = barcode_info.split("^")
                    detect_num = len(barcodes)
                    self.dark_detect_nums.append(detect_num)
                    decode_num = 0
                    for b in barcodes:
                        decode_info = b.split(",")[0]
                        if decode_info != "":
                            decode_num += 1
                    self.dark_decode_nums.append(decode_num)
                    decode_rate = decode_num/detect_num
                    self.dark_decode_rates.append(decode_rate)
            if sum(self.dark_detect_nums):
                self.dark_paths.append("-")
                self.dark_detect_nums.append(sum(self.dark_detect_nums))
                self.dark_decode_nums.append(sum(self.dark_decode_nums))
                self.dark_decode_rates.append(sum(self.dark_decode_nums)/sum(self.dark_detect_nums))
                self.dark_spend_times.append("-")
                self.dark_scene.append("dark total")

             # overexposure  scene total
            self.overexposure_paths = []
            self.overexposure_detect_nums = []
            self.overexposure_decode_nums = []
            self.overexposure_decode_rates = []
            self.overexposure_spend_times = []
            self.overexposure_scene = []
            for line in self.txt_lines:
                file_path,barcode_info,latency = line.split(" ")
                if "overexposure" in file_path:
                    self.overexposure_paths.append(file_path)
                    self.overexposure_scene.append("overexposure")
                    self.overexposure_spend_times.append(latency)
                    barcodes = barcode_info.split("^")
                    detect_num = len(barcodes)
                    self.overexposure_detect_nums.append(detect_num)
                    decode_num = 0
                    for b in barcodes:
                        decode_info = b.split(",")[0]
                        if decode_info != "":
                            decode_num += 1
                    self.overexposure_decode_nums.append(decode_num)
                    decode_rate = decode_num/detect_num
                    self.overexposure_decode_rates.append(decode_rate)

            if sum(self.overexposure_detect_nums):
                self.overexposure_paths.append("-")
                self.overexposure_detect_nums.append(sum(self.overexposure_detect_nums))
                self.overexposure_decode_nums.append(sum(self.overexposure_decode_nums))
                self.overexposure_decode_rates.append(sum(self.overexposure_decode_nums)/sum(self.overexposure_detect_nums))
                self.overexposure_spend_times.append("-")
                self.overexposure_scene.append("overexposure total")
            
            # 倾斜条码信息
            self.tilted_paths = []
            self.tilted_detect_nums = []
            self.tilted_decode_nums = []
            self.tilted_decode_rates = []
            self.tilted_spend_times = []
            self.tilted_scene = []
            for line in self.txt_lines:
                file_path,barcode_info,latency = line.split(" ")
                if "tilted" in file_path:
                    self.tilted_paths.append(file_path)
                    self.tilted_scene.append("tilted")
                    self.tilted_spend_times.append(latency)
                    barcodes = barcode_info.split("^")
                    detect_num = len(barcodes)
                    self.tilted_detect_nums.append(detect_num)
                    decode_num = 0
                    for b in barcodes:
                        decode_info = b.split(",")[0]
                        if decode_info != "":
                            decode_num += 1
                    self.tilted_decode_nums.append(decode_num)
                    decode_rate = decode_num/detect_num
                    self.tilted_decode_rates.append(decode_rate)
            if sum(self.tilted_detect_nums):
                self.tilted_paths.append("-")
                self.tilted_detect_nums.append(sum(self.tilted_detect_nums))
                self.tilted_decode_nums.append(sum(self.tilted_decode_nums))
                self.tilted_decode_rates.append(sum(self.tilted_decode_nums)/sum(self.tilted_detect_nums))
                self.tilted_spend_times.append("-")
                self.tilted_scene.append("tilted total")
            
            # 汇总所有信息
            self.paths += self.dark_paths + self.overexposure_paths + self.tilted_paths
            self.detect_nums += self.dark_detect_nums + self.overexposure_detect_nums + self.tilted_detect_nums
            self.decode_nums += self.dark_decode_nums + self.overexposure_decode_nums + self.tilted_decode_nums
            self.decode_rates += self.dark_decode_rates + self.overexposure_decode_rates + self.tilted_decode_rates
            self.spend_times += self.dark_spend_times + self.overexposure_spend_times + self.tilted_spend_times
            self.scene += self.dark_scene + self.overexposure_scene + self.tilted_scene
            
            if sum(self.detect_nums):
                self.paths.append("-")
                self.detect_nums.append(sum(self.detect_nums))
                self.decode_nums.append(sum(self.decode_nums))
                self.decode_rates.append(sum(self.decode_nums)/sum(self.detect_nums))
                self.spend_times.append("-")
                self.scene.append("total")

            self.result["scene"] = self.scene
            self.result["path"] = self.paths
            self.result["detect nums"] = self.detect_nums
            self.result["decode nums"] = self.decode_nums
            self.result["decode rates"] = self.decode_rates
            self.result["spend time"] = self.spend_times

    def save_to_excel(self):
        df = pd.DataFrame(self.result)        
        df.to_excel(self.result_path, index=False)
        workbook = load_workbook(self.result_path)
        worksheet = workbook.active
        highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        keyword = "total"
        # 遍历DataFrame的每一行，查找包含关键字的行并高亮
        for index, row in df.iterrows():
            if keyword in str(row.values):
                
                # 获取该行在工作表中的数据范围
                for col_num in range(1, 7):  
                    cell = worksheet.cell(row=index + 2, column=col_num)
                    cell.fill = highlight_fill
        workbook.save(self.result_path)
    
    def run(self):
        self.read()
        self.save_to_excel()


if __name__=="__main__":
    result_txt = "result.txt"
    save_excel = "result.xlsx"
    rd = ResultDecoder(result_txt,save_excel)
    rd.run()
                
    
